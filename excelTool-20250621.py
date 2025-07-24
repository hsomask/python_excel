# Auther lzm

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font, Border, Side
import os
import logging
import xlrd
from itertools import groupby
from operator import itemgetter
import inspect
from datetime import datetime


# 设置日志
import sys
import tempfile

# 优先尝试脚本目录
script_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.path.dirname(sys.argv[0])
log_dir = os.path.join(script_dir, "LOG")
log_file = os.path.join(log_dir, "excel_tool.log")

# 如果是打包后的EXE，使用与EXE同级的LOG目录
if getattr(sys, 'frozen', False):
    exe_dir = os.path.dirname(sys.executable)
    log_dir = os.path.join(exe_dir, "LOG")
    log_file = os.path.join(log_dir, "excel_tool.log")

print(f"1. 日志文件路径: {log_file}")
print(f"2. 运行模式: {'EXE' if getattr(sys, 'frozen', False) else '脚本'}")
print(f"3. 最终日志目录: {log_dir}")

# 确保日志目录存在
try:
    if not os.path.exists(log_dir):
        print("2. 尝试创建日志目录...")
        os.makedirs(log_dir)
        print(f"3. 成功创建日志目录: {log_dir}")
    else:
        print(f"3. 日志目录已存在: {log_dir}")

    # 测试目录可写性
    test_file = os.path.join(log_dir, "test_write.tmp")
    with open(test_file, 'w') as f:
        f.write("test")
    os.remove(test_file)
    print("4. 目录可写性测试通过")

except Exception as e:
    print(f"5. 脚本目录日志创建失败: {e}")
    log_dir = os.path.join(os.path.expanduser("~"), "excel_tool_logs")
    log_file = os.path.join(log_dir, "excel_tool.log")
    print(f"6. 尝试在用户目录创建日志: {log_file}")
    try:
        os.makedirs(log_dir, exist_ok=True)
        print(f"7. 用户目录日志准备完成: {log_dir}")
    except Exception as e:
        print(f"8. 用户目录日志创建失败: {e}")
        log_dir = os.path.join(os.getcwd(), "temp_logs")
        log_file = os.path.join(log_dir, "excel_tool.log")
        print(f"9. 最后尝试在当前目录创建日志: {log_file}")
        os.makedirs(log_dir, exist_ok=True)

# 配置日志
print(f"10. 准备配置日志到: {log_file}")
try:
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s (Line: %(lineno)d)',
        handlers=[
            logging.FileHandler(log_file, mode='a', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger(__name__)
    logger.info("=" * 50)
    logger.info("日志系统初始化成功")
    logger.info(f"日志文件位置: {log_file}")
    logger.info("=" * 50)
    print(f"11. 日志系统已初始化，日志文件: {log_file}")

    # 确认日志文件已创建
    if os.path.exists(log_file):
        print(f"12. 确认日志文件存在: {log_file}")
        print(f"13. 日志文件大小: {os.path.getsize(log_file)} 字节")
    else:
        print(f"12. 警告: 日志文件不存在: {log_file}")

except Exception as e:
    print(f"14. 日志配置失败: {e}")
    print("15. 尝试简单日志配置...")
    try:
        # logging.basicConfig(
        #     level=logging.INFO,
        #     format='%(asctime)s [%(levelname)s] %(message)s',
        #     handlers=[logging.StreamHandler()]
        # )
        logger = logging.getLogger(__name__)
        logger.error(f"无法创建文件日志，仅使用控制台日志。原错误: {e}")
    except Exception as e:
        print(f"16. 简单日志配置也失败: {e}")
        raise

def expand_month_range(start, end):
    result = []
    current = datetime.strptime(start, "%Y%m")
    end_date = datetime.strptime(end, "%Y%m")
    while current <= end_date:
        result.append(current.strftime("%Y.%m"))
        # 增加一个月
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return result

def merge_consecutive_dates(dates):
    """合并连续的日期"""
    # 将日期转换为可比较的元组 (年, 月)
    date_tuples = []
    for date in dates:
        year, month = map(int, date.split('.'))
        date_tuples.append((year, month, date))  # 保留原始字符串格式

    # 按年月排序
    date_tuples.sort()

    ranges = []
    current_range = []

    for i in range(len(date_tuples)):
        if not current_range:
            current_range.append(date_tuples[i])
            continue

        last_year, last_month, _ = current_range[-1]
        current_year, current_month, _ = date_tuples[i]

        # 检查是否连续 (相差1个月)
        if (current_year == last_year and current_month == last_month + 1) or \
                (current_year == last_year + 1 and current_month == 1 and last_month == 12):
            current_range.append(date_tuples[i])
        else:
            # 结束当前范围
            if len(current_range) > 1:
                start = current_range[0][2]  # 原始字符串
                end = current_range[-1][2]  # 原始字符串
                ranges.append(f"{start}-{end}")
            else:
                ranges.append(current_range[0][2])
            current_range = [date_tuples[i]]

    # 处理最后一个范围
    if current_range:
        if len(current_range) > 1:
            start = current_range[0][2]
            end = current_range[-1][2]
            ranges.append(f"{start}-{end}")
        else:
            ranges.append(current_range[0][2])

    return ranges


def process_extract_records(extract_df, deposit_date_ranges):
    """处理提取记录数据"""
    extract_records = []
    valid_rows = extract_df.iloc[7:]  # 从第8行开始

    logger.info(f"Processing {len(valid_rows)} extract records")
    logger.info(f"Deposit date ranges: {deposit_date_ranges}")

    for idx, row in valid_rows.iterrows():
        try:
            company = str(row[1]) if pd.notna(row[1]) else ""  # B列单位名称
            date_str = str(row[3]) if pd.notna(row[3]) else ""  # D列业务时间
            transaction_status = str(row[10]) if pd.notna(row[10]) else ""  # k列交易状态
            logger.info(f"Processing row {idx}: date_str={date_str} (type: {type(date_str)})")
            
            if transaction_status != "正常":
                logger.info(f"Skipping row {idx} due to non-normal transaction status: {transaction_status}")
                continue
            # 处理多种日期格式为YYYY.MM.DD
            if len(date_str) == 6 and date_str.isdigit():  # 格式为yyyymm
                date = f"{date_str[:4]}.{date_str[4:6]}.01"  # 转为yyyy.mm.dd
                logger.info(f"Converted date format: {date}")
            elif len(date_str) >= 10 and (date_str[4] == '-' or date_str[4] == '/'):  # 格式为yyyy-mm-dd或yyyy/mm/dd
                date = f"{date_str[:4]}.{date_str[5:7]}.{date_str[8:10]}"  # 转为yyyy.mm.dd
                logger.info(f"Converted date format: {date}")
            else:
                logger.info(f"Keeping original date format: {date_str}")
                date = date_str  # 保持原始格式

            extract_type = str(row[7]) if pd.notna(row[7]) else ""  # H列提取类型
            amount = float(row[8]) if pd.notna(row[8]) else 0.0  # I列提取金额

            # 找到对应的缴存时间段
            matched = False
            for d_range in deposit_date_ranges:
                parts = d_range.split('-')
                start = parts[0]
                end = parts[-1] if len(parts) > 1 else parts[0]

                logger.info(f"Checking range {d_range} (start={start}, end={end}) against date {date}")

                if start <= date <= end:
                    logger.info(f"Match found for date {date} in range {d_range}")
                    extract_records.append({
                        'deposit_range': d_range,
                        'company': company,
                        'date': date,
                        'type': extract_type,
                        'amount': amount
                    })
                    matched = True
                    break
                elif  transaction_status == "正常":
                    # 如果不在任何缴存区间内，但状态为正常，也记录提取信息
                    logger.info(f"Recording normal extract outside deposit range: {date}")
                    extract_records.append({
                        'deposit_range': d_range,  # 标记为未匹配
                        'company': company,
                        'date': date,
                        'type': extract_type,
                        'amount': amount
                    })
                    matched = True
                    break

            if not matched:
                logger.info(f"No matching deposit range found for date {date}")
        except Exception as e:
            logger.error(f"Error processing extract record: {e}")
            continue

    # 合并连续且类型相同的记录
    merged_records = []
    # 按照提取日期排序(从小到大)，然后按缴存时间段和类型排序
    extract_records.sort(key=lambda x: (
        float(x['date'].replace('.', '')),  # 主排序键：提取日期
        float(x['deposit_range'].split('-')[0].replace('.', '')),  # 次排序键：缴存时间段起始日期
        x['type']  # 最后按类型排序
    ))

    current_record = None
    for record in extract_records:
        if not current_record:
            current_record = record.copy()
            continue

        # 检查是否可以合并
        last_date = current_record['date'].split('.')
        curr_date = record['date'].split('.')
        same_type = current_record['type'] == record['type']
        same_range = current_record['deposit_range'] == record['deposit_range']

        # 计算月份差
        month_diff = (int(curr_date[0]) - int(last_date[0])) * 12 + (int(curr_date[1]) - int(last_date[1]))

        if same_type and same_range and month_diff == 1:
            # 合并记录
            current_record['amount'] += record['amount']
            current_record['end_date'] = record['date']
        elif same_type and same_range and month_diff > 1 and 'end_date' in current_record:
            # 检查是否与已有区间连续
            last_end_date = current_record['end_date'].split('.')
            curr_end_diff = (int(curr_date[0]) - int(last_end_date[0])) * 12 + (
                        int(curr_date[1]) - int(last_end_date[1]))
            if curr_end_diff == 1:
                current_record['amount'] += record['amount']
                current_record['end_date'] = record['date']
        else:
            # 保存当前记录并开始新记录
            if 'end_date' in current_record:
                # 日期区间格式化为yyyy.mm-yyyy.mm (去掉日期部分)
                start = '.'.join(current_record['date'].split('.')[:2])  # 取yyyy.mm
                end = '.'.join(current_record['end_date'].split('.')[:2])  # 取yyyy.mm
                current_record['date_range'] = f"{start}-{end}"
            else:
                # 保持原始日期格式不变
                current_record['date_range'] = current_record['date']
            merged_records.append(current_record)
            current_record = record.copy()

    # 添加最后一个记录
    if current_record:
        if 'end_date' in current_record:
            # 日期区间格式化为yyyy.mm-yyyy.mm (去掉日期部分)
            start = '.'.join(current_record['date'].split('.')[:2])  # 取yyyy.mm
            end = '.'.join(current_record['end_date'].split('.')[:2])  # 取yyyy.mm
            current_record['date_range'] = f"{start}-{end}"
        else:
            current_record['date_range'] = current_record['date']
        merged_records.append(current_record)

    return merged_records


# 读取缴存.xls文件
# with pd.ExcelFile("./DATA/DepositDetails.xls") as excel:
with pd.ExcelFile("../DATA/DepositDetails.xls") as excel:
    df = pd.read_excel(excel, header=None)
    name = df.iloc[0, 6]  # G1单元格
    id_num = df.iloc[0, 9]  # J1单元格
    grzhye = df[df[5].notna()].iloc[-1, 5]  # F列最后一个非空值

    # 处理缴存记录数据
    jc_records = []
    valid_types = ["汇缴", "补缴", "补缴往月", "少缴补缴", "差额补缴"]
    data_rows = df.iloc[5:]  # 从第6行开始

    # 筛选有效记录并按单位分组
    records = []
    # lzm 20250621优化少缴补缴月份
    for _, row in data_rows.iterrows():
        date_str = str(row[9])
        if row[1] in valid_types:
            if "-" in date_str:
                start, end = date_str.split("-")
                date_range = expand_month_range(start, end)
                for date in date_range:
                    records.append({
                        'company': row[7],
                        'date': date
                    })
            else:
            # if row[1] in valid_types:  # B列为业务类型
                records.append({
                    'company': row[7],  # H列为单位名称
                    'date': str(row[9])[:4] + "." + str(row[9])[4:6]  # J列为时间，格式化为yyyy.mm
                })

    # 按单位分组并合并日期
    company_groups = {}
    for record in records:
        if record['company'] not in company_groups:
            company_groups[record['company']] = []
        company_groups[record['company']].append(record['date'])

    # 合并每个单位的连续日期
    for company, dates in company_groups.items():
        merged_dates = merge_consecutive_dates(list(set(dates)))  # 去重并合并
        for date_range in merged_dates:
            jc_records.append({
                'company': company,
                'date_range': date_range,
                'sort_key': float(date_range.split('-')[0].replace('.', ''))  # 提取最早时间作为排序依据
            })

    # 按时间排序记录 - 确保按缴存时间段的起始日期正确排序
    jc_records.sort(key=lambda x: float(x['date_range'].split('-')[0].replace('.', '')))



####-----------------------lzm begin--------------------------------
def process_deposit_sheet(df):
    """处理单个sheet的缴存记录"""
    try:
        # 获取基本信息
        name = str(df.iloc[0, 6]) if pd.notna(df.iloc[0, 6]) else ""  # G1单元格
        id_num = str(df.iloc[0, 9]) if pd.notna(df.iloc[0, 9]) else ""  # J1单元格

        # 获取账户余额（F列最后一个非空值）
        grzhye = 0.0
        balance_rows = df[df[5].notna()]  # F列非空行
        if not balance_rows.empty:
            grzhye = float(balance_rows.iloc[-1, 5])

        # 处理缴存记录数据
        jc_records = []
        valid_types = ["汇缴", "补缴", "补缴往月", "少缴补缴", "差额补缴"]
        data_rows = df.iloc[5:]  # 从第6行开始

        # 筛选有效记录并按单位分组
        records = []
        for _, row in data_rows.iterrows():
            date_str = str(row[9])
            if row[1] in valid_types:
                if "-" in date_str:
                    start, end = date_str.split("-")
                    date_range = expand_month_range(start, end)
                    for date in date_range:
                        records.append({
                            'company': row[7],
                            'date': date
                        })
            else:
                # if row[1] in valid_types:  # B列为业务类型
                records.append({
                    'company': row[7],  # H列为单位名称
                    'date': str(row[9])[:4] + "." + str(row[9])[4:6]  # J列为时间，格式化为yyyy.mm
                })
            # if pd.notna(row[1]) and str(row[1]) in valid_types:  # B列为业务类型
            #     try:
            #         date_str = str(row[9])  # J列为时间
            #         if len(date_str) >= 6:  # 确保日期格式正确
            #             date = date_str[:4] + "." + date_str[4:6]  # 格式化为yyyy.mm
            #             records.append({
            #                 'company': str(row[7]) if pd.notna(row[7]) else "",  # H列为单位名称
            #                 'date': date,
            #                 'base': float(row[2]) if pd.notna(row[2]) else 0.0,  # C列为缴存基数
            #                 'amount': float(row[4]) if pd.notna(row[4]) else 0.0  # F列为月缴存额
            #             })


        # 按单位分组并合并日期
        company_groups = {}
        for record in records:
            if record['company'] not in company_groups:
                company_groups[record['company']] = {
                    'dates': [],
                    'base': record['base'],
                    'amount': record['amount']
                }
            company_groups[record['company']]['dates'].append(record['date'])

        # 合并每个单位的连续日期
        for company, data in company_groups.items():
            merged_dates = merge_consecutive_dates(list(set(data['dates'])))  # 去重并合并
            for date_range in merged_dates:
                jc_records.append({
                    'company': company,
                    'date_range': date_range,
                    'base': data['base'],
                    'amount': data['amount'],
                    'sort_key': float(date_range.split('-')[0].replace('.', ''))
                })

        # 按时间排序记录
        jc_records.sort(key=lambda x: x['sort_key'])

        return {
            'name': name,
            'id_num': id_num,
            'grzhye': grzhye,
            'jc_records': jc_records
        }

    except Exception as e:
        logger.error(f"处理缴存记录sheet时出错: {str(e)}")
        logger.error(f"错误详情: {repr(e)}")
        return {
            'name': "",
            'id_num': "",
            'grzhye': 0.0,
            'jc_records': []
        }
def process_extract_sheet(df, deposit_ranges):
    """处理单个sheet的提取记录"""
    return process_extract_records(df, deposit_ranges)


def create_excel_sheet(new_wb, sheet_name, person_data, extract_records):
    """为单个人创建Excel sheet"""
    new_ws = new_wb.create_sheet(title=sheet_name)

    # 写入基础数据
    new_ws['B5'] = person_data['name']
    new_ws['C5'] = person_data['id_num']
    new_ws['D5'] = person_data['grzhye']

    # 写入缴存记录和对应的提取记录
    current_row = 5
    for record in person_data['jc_records']:
        # 写入缴存记录(只写一次)
        new_ws[f'E{current_row}'] = record['date_range']  # 写入时间段到E列
        new_ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        new_ws[f'F{current_row}'] = record['company']  # 写入单位到F列
        new_ws[f'F{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # 写入对应的提取记录
        matching_extracts = [e for e in extract_records if e['deposit_range'] == record['date_range']]
        if matching_extracts:
            # 合并缴存记录单元格
            end_row = current_row + len(matching_extracts) - 1
            if end_row > current_row:
                new_ws.merge_cells(f'E{current_row}:E{end_row}')
                new_ws.merge_cells(f'F{current_row}:F{end_row}')

            for ext in matching_extracts:
                new_ws[f'I{current_row}'] = ext['date_range']  # 提取时间段
                new_ws[f'I{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                new_ws[f'J{current_row}'] = ext['amount']  # 提取金额
                new_ws[f'J{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                new_ws[f'K{current_row}'] = ext['type']  # 提取类型
                new_ws[f'K{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
        else:
            # 即使没有提取记录也要写入缴存记录
            new_ws[f'I{current_row}'] = ""
            new_ws[f'J{current_row}'] = ""
            new_ws[f'K{current_row}'] = ""
            current_row += 1

    # 合并单元格并设置居中
    last_row = current_row - 1
    if last_row > 5:
        # 合并姓名列(B列)
        new_ws.merge_cells(f'B5:B{last_row}')
        new_ws['B5'].alignment = Alignment(horizontal='center', vertical='center')

        # 合并序号列(A列)
        new_ws.merge_cells(f'A5:A{last_row}')
        new_ws['A5'].alignment = Alignment(horizontal='center', vertical='center')

        # 合并证件号码列(C列)
        new_ws.merge_cells(f'C5:C{last_row}')
        new_ws['C5'].alignment = Alignment(horizontal='center', vertical='center')

        # 合并账户余额列(D列)
        new_ws.merge_cells(f'D5:D{last_row}')
        new_ws['D5'].alignment = Alignment(horizontal='center', vertical='center')

        # 合并当前缴存基数列(G列)
        new_ws.merge_cells(f'G5:G{last_row}')
        new_ws['G5'].alignment = Alignment(horizontal='center', vertical='center')

        # 合并当前月缴存额列(H列)
        new_ws.merge_cells(f'H5:H{last_row}')
        new_ws['H5'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置表头样式
    new_ws.merge_cells('A1:K1')
    new_ws['A1'].value = "附件1"
    new_ws['A1'].font = Font(name='宋体', size=20, bold=True)
    new_ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    new_ws.merge_cells('A2:K2')
    new_ws['A2'].value = "住房公积金基本信息"
    new_ws['A2'].font = Font(name='宋体', size=20, bold=True)
    new_ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置列标题
    headers = {
        'A3': '序号', 'B3': '姓名', 'C3': '证件号码', 'D3': '账户余额',
        'E3': '缴存记录', 'I3': '提取记录'
    }
    for cell, value in headers.items():
        new_ws[cell].value = value
        new_ws[cell].font = Font(name='宋体', size=11, bold=True)
        new_ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # 设置子标题
    sub_headers = {
        'E4': '缴存时间段', 'F4': '公积金缴存单位', 'G4': '当前缴存基数',
        'H4': '当前月缴存额', 'I4': '提取时间', 'J4': '提取金额', 'K4': '提取类型'
    }
    for cell, value in sub_headers.items():
        new_ws[cell].value = value
        new_ws[cell].font = Font(name='宋体', size=11, bold=True)
        new_ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    column_widths = {
        'A': 13, 'B': 13, 'C': 20, 'D': 13, 'E': 20,
        'F': 35, 'G': 13, 'H': 13, 'I': 20, 'J': 13, 'K': 13
    }
    for col, width in column_widths.items():
        new_ws.column_dimensions[col].width = width

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in new_ws.iter_rows(min_row=1, max_row=new_ws.max_row):
        for cell in row:
            cell.border = thin_border

    # 添加查询截止时间行
    last_row = new_ws.max_row + 1
    new_ws.merge_cells(f'A{last_row}:K{last_row}')
    new_ws[f'A{last_row}'].value = "查询截止时间为YYYY年MM月DD日HH时MI分"
    new_ws[f'A{last_row}'].font = Font(name='宋体', size=12)
    new_ws[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')

    return new_ws
def main():
    data_dir = "../DATA"
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
        logger.info(f"Created directory: {data_dir}")

    output_path = os.path.join(data_dir, "output.xlsx")
    logger.info(f"Attempting to save to: {output_path}")
    # 创建新的Excel工作簿
    print(os.getcwd())

    new_wb = Workbook()
    # 删除默认创建的sheet
    new_wb.remove(new_wb.active)

    # 读取缴存.xls文件
    
    deposit_records = {}
    try:
        with pd.ExcelFile("../DATA/DepositDetails.xls") as excel:
            for sheet_name in excel.sheet_names:
                try:
                    df = pd.read_excel(excel, sheet_name=sheet_name, header=None)
                    person_data = process_deposit_sheet(df)
                    deposit_records[sheet_name] = person_data
                    logger.info(f"成功处理缴存记录sheet: {sheet_name}")
                except Exception as e:
                    logger.error(f"处理缴存记录sheet {sheet_name} 时出错: {str(e)}")
                    continue
    except Exception as e:
        logger.error(f"读取缴存文件时出错: {str(e)}")
        raise

    # 读取提取.xls文件
    extract_records = {}
    extract_file_path = "../DATA/ExtractDetails.xls"

    if os.path.exists(extract_file_path):
        try:
            with pd.ExcelFile(extract_file_path) as excel:
                # 获取提取文件中的所有sheet名称
                extract_sheets = set(excel.sheet_names)
                logger.info(f"提取文件中包含的sheet: {extract_sheets}")
                
                # 获取缴存文件中的所有sheet名称
                deposit_sheets = set(deposit_records.keys())
                logger.info(f"缴存文件中包含的sheet: {deposit_sheets}")
                
                # 找出两个文件共有的sheet
                common_sheets = deposit_sheets & extract_sheets
                logger.info(f"两个文件共有的sheet: {common_sheets}")
                
                # 处理共有的sheet
                for sheet_name in common_sheets:
                    try:
                        df = pd.read_excel(excel, sheet_name=sheet_name, header=None)
                        deposit_ranges = [r['date_range'] for r in deposit_records[sheet_name]['jc_records']]
                        extract_records[sheet_name] = process_extract_sheet(df, deposit_ranges)
                        logger.info(f"成功处理提取记录sheet: {sheet_name}")
                    except Exception as e:
                        logger.error(f"处理提取记录sheet {sheet_name} 时出错: {str(e)}")
                        extract_records[sheet_name] = []
                
                # 记录只在缴存文件中存在的sheet
                deposit_only_sheets = deposit_sheets - extract_sheets
                if deposit_only_sheets:
                    logger.info(f"以下sheet只在缴存文件中存在: {deposit_only_sheets}")
                    for sheet_name in deposit_only_sheets:
                        extract_records[sheet_name] = []
                
                # 记录只在提取文件中存在的sheet
                extract_only_sheets = extract_sheets - deposit_sheets
                if extract_only_sheets:
                    logger.info(f"以下sheet只在提取文件中存在: {extract_only_sheets}")
        except Exception as e:
            logger.warning(f"读取提取文件时出错，将继续处理缴存记录: {str(e)}")
    else:
        logger.warning("未找到提取文件，将仅处理缴存记录")
        # 为每个缴存记录初始化空的提取记录
        for sheet_name in deposit_records:
            extract_records[sheet_name] = []


    # 为每个人创建Excel sheet
    for sheet_name, person_data in deposit_records.items():
        try:
            person_extract_records = extract_records.get(sheet_name, [])
            create_excel_sheet(new_wb, sheet_name, person_data, person_extract_records)
            logger.info(f"成功创建sheet: {sheet_name}")
        except Exception as e:
            logger.error(f"创建sheet {sheet_name} 时出错: {str(e)}")
            continue

    # 保存文件
    output_path = os.path.join(data_dir, "output.xlsx")
    try:
        if os.path.exists(output_path):
            os.remove(output_path)
        new_wb.save(output_path)
        logger.info(f"成功保存文件到: {output_path}")
    except Exception as e:
        logger.error(f"保存文件时出错: {str(e)}")
        raise

if __name__ == "__main__":
    main()
####-----------------------lzm end----------------------------------




