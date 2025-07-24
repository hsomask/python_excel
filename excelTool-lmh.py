# Auther lmh

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font, Border, Side
import os
import logging
from itertools import groupby
from operator import itemgetter
import inspect

# 设置日志
import sys
import tempfile

# 优先尝试脚本目录
script_dir = os.path.dirname(os.path.abspath(__file__)) if __file__ else os.path.dirname(sys.argv[0])
log_dir = os.path.join(script_dir, "LOG")
log_file = os.path.join(log_dir, "excel_tool.log")

# 如果是打包后的EXE，使用与EXE同级的LOG目录
if getattr(sys, 'frozen', False):
    exe_dir = os.path.dirname(sys.executable)
    log_dir = os.path.join(exe_dir, "LOG")
    log_file = os.path.join(log_dir, "excel_tool.log")

print(f"1. 日志文件路径: {log_file}")
print(f"2. 运行模式: {'EXE' if getattr(sys, 'frozen', False) else '脚本'}")
print(f"3. 最终日志目录: {log_dir}")

# 确保日志目录存在
try:
    if not os.path.exists(log_dir):
        print("2. 尝试创建日志目录...")
        os.makedirs(log_dir)
        print(f"3. 成功创建日志目录: {log_dir}")
    else:
        print(f"3. 日志目录已存在: {log_dir}")

    # 测试目录可写性
    test_file = os.path.join(log_dir, "test_write.tmp")
    with open(test_file, 'w') as f:
        f.write("test")
    os.remove(test_file)
    print("4. 目录可写性测试通过")

except Exception as e:
    print(f"5. 脚本目录日志创建失败: {e}")
    log_dir = os.path.join(os.path.expanduser("~"), "excel_tool_logs")
    log_file = os.path.join(log_dir, "excel_tool.log")
    print(f"6. 尝试在用户目录创建日志: {log_file}")
    try:
        os.makedirs(log_dir, exist_ok=True)
        print(f"7. 用户目录日志准备完成: {log_dir}")
    except Exception as e:
        print(f"8. 用户目录日志创建失败: {e}")
        log_dir = os.path.join(os.getcwd(), "temp_logs")
        log_file = os.path.join(log_dir, "excel_tool.log")
        print(f"9. 最后尝试在当前目录创建日志: {log_file}")
        os.makedirs(log_dir, exist_ok=True)

# 配置日志
print(f"10. 准备配置日志到: {log_file}")
try:
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s (Line: %(lineno)d)',
        handlers=[
            logging.FileHandler(log_file, mode='a', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    logger = logging.getLogger(__name__)
    logger.info("=" * 50)
    logger.info("日志系统初始化成功")
    logger.info(f"日志文件位置: {log_file}")
    logger.info("=" * 50)
    print(f"11. 日志系统已初始化，日志文件: {log_file}")

    # 确认日志文件已创建
    if os.path.exists(log_file):
        print(f"12. 确认日志文件存在: {log_file}")
        print(f"13. 日志文件大小: {os.path.getsize(log_file)} 字节")
    else:
        print(f"12. 警告: 日志文件不存在: {log_file}")

except Exception as e:
    print(f"14. 日志配置失败: {e}")
    print("15. 尝试简单日志配置...")
    try:
        # logging.basicConfig(
        #     level=logging.INFO,
        #     format='%(asctime)s [%(levelname)s] %(message)s',
        #     handlers=[logging.StreamHandler()]
        # )
        logger = logging.getLogger(__name__)
        logger.error(f"无法创建文件日志，仅使用控制台日志。原错误: {e}")
    except Exception as e:
        print(f"16. 简单日志配置也失败: {e}")
        raise


def merge_consecutive_dates(dates):
    """合并连续的日期"""
    # 将日期转换为可比较的元组 (年, 月)
    date_tuples = []
    for date in dates:
        year, month = map(int, date.split('.'))
        date_tuples.append((year, month, date))  # 保留原始字符串格式

    # 按年月排序
    date_tuples.sort()

    ranges = []
    current_range = []

    for i in range(len(date_tuples)):
        if not current_range:
            current_range.append(date_tuples[i])
            continue

        last_year, last_month, _ = current_range[-1]
        current_year, current_month, _ = date_tuples[i]

        # 检查是否连续 (相差1个月)
        if (current_year == last_year and current_month == last_month + 1) or \
                (current_year == last_year + 1 and current_month == 1 and last_month == 12):
            current_range.append(date_tuples[i])
        else:
            # 结束当前范围
            if len(current_range) > 1:
                start = current_range[0][2]  # 原始字符串
                end = current_range[-1][2]  # 原始字符串
                ranges.append(f"{start}-{end}")
            else:
                ranges.append(current_range[0][2])
            current_range = [date_tuples[i]]

    # 处理最后一个范围
    if current_range:
        if len(current_range) > 1:
            start = current_range[0][2]
            end = current_range[-1][2]
            ranges.append(f"{start}-{end}")
        else:
            ranges.append(current_range[0][2])

    return ranges


def process_extract_records(extract_df, deposit_date_ranges):
    """处理提取记录数据"""
    extract_records = []
    valid_rows = extract_df.iloc[7:]  # 从第8行开始

    logger.info(f"Processing {len(valid_rows)} extract records")
    logger.info(f"Deposit date ranges: {deposit_date_ranges}")

    for idx, row in valid_rows.iterrows():
        try:
            company = str(row[1]) if pd.notna(row[1]) else ""  # B列单位名称
            date_str = str(row[3]) if pd.notna(row[3]) else ""  # D列业务时间
            logger.info(f"Processing row {idx}: date_str={date_str} (type: {type(date_str)})")

            # 处理多种日期格式为YYYY.MM.DD
            if len(date_str) == 6 and date_str.isdigit():  # 格式为yyyymm
                date = f"{date_str[:4]}.{date_str[4:6]}.01"  # 转为yyyy.mm.dd
                logger.info(f"Converted date format: {date}")
            elif len(date_str) >= 10 and (date_str[4] == '-' or date_str[4] == '/'):  # 格式为yyyy-mm-dd或yyyy/mm/dd
                date = f"{date_str[:4]}.{date_str[5:7]}.{date_str[8:10]}"  # 转为yyyy.mm.dd
                logger.info(f"Converted date format: {date}")
            else:
                logger.info(f"Keeping original date format: {date_str}")
                date = date_str  # 保持原始格式

            extract_type = str(row[7]) if pd.notna(row[7]) else ""  # H列提取类型
            amount = float(row[8]) if pd.notna(row[8]) else 0.0  # I列提取金额

            # 找到对应的缴存时间段
            matched = False
            for d_range in deposit_date_ranges:
                parts = d_range.split('-')
                start = parts[0]
                end = parts[-1] if len(parts) > 1 else parts[0]

                logger.info(f"Checking range {d_range} (start={start}, end={end}) against date {date}")

                if start <= date <= end:
                    logger.info(f"Match found for date {date} in range {d_range}")
                    extract_records.append({
                        'deposit_range': d_range,
                        'company': company,
                        'date': date,
                        'type': extract_type,
                        'amount': amount
                    })
                    matched = True
                    break

            if not matched:
                logger.info(f"No matching deposit range found for date {date}")
        except Exception as e:
            logger.error(f"Error processing extract record: {e}")
            continue

    # 合并连续且类型相同的记录
    merged_records = []
    # 按照提取日期排序(从小到大)，然后按缴存时间段和类型排序
    extract_records.sort(key=lambda x: (
        float(x['date'].replace('.', '')),  # 主排序键：提取日期
        float(x['deposit_range'].split('-')[0].replace('.', '')),  # 次排序键：缴存时间段起始日期
        x['type']  # 最后按类型排序
    ))

    current_record = None
    for record in extract_records:
        if not current_record:
            current_record = record.copy()
            continue

        # 检查是否可以合并
        last_date = current_record['date'].split('.')
        curr_date = record['date'].split('.')
        same_type = current_record['type'] == record['type']
        same_range = current_record['deposit_range'] == record['deposit_range']

        # 计算月份差
        month_diff = (int(curr_date[0]) - int(last_date[0])) * 12 + (int(curr_date[1]) - int(last_date[1]))

        if same_type and same_range and month_diff == 1:
            # 合并记录
            current_record['amount'] += record['amount']
            current_record['end_date'] = record['date']
        elif same_type and same_range and month_diff > 1 and 'end_date' in current_record:
            # 检查是否与已有区间连续
            last_end_date = current_record['end_date'].split('.')
            curr_end_diff = (int(curr_date[0]) - int(last_end_date[0])) * 12 + (
                        int(curr_date[1]) - int(last_end_date[1]))
            if curr_end_diff == 1:
                current_record['amount'] += record['amount']
                current_record['end_date'] = record['date']
        else:
            # 保存当前记录并开始新记录
            if 'end_date' in current_record:
                # 日期区间格式化为yyyy.mm-yyyy.mm (去掉日期部分)
                start = '.'.join(current_record['date'].split('.')[:2])  # 取yyyy.mm
                end = '.'.join(current_record['end_date'].split('.')[:2])  # 取yyyy.mm
                current_record['date_range'] = f"{start}-{end}"
            else:
                # 保持原始日期格式不变
                current_record['date_range'] = current_record['date']
            merged_records.append(current_record)
            current_record = record.copy()

    # 添加最后一个记录
    if current_record:
        if 'end_date' in current_record:
            # 日期区间格式化为yyyy.mm-yyyy.mm (去掉日期部分)
            start = '.'.join(current_record['date'].split('.')[:2])  # 取yyyy.mm
            end = '.'.join(current_record['end_date'].split('.')[:2])  # 取yyyy.mm
            current_record['date_range'] = f"{start}-{end}"
        else:
            current_record['date_range'] = current_record['date']
        merged_records.append(current_record)

    return merged_records


# 读取缴存.xls文件
with pd.ExcelFile("./DATA/DepositDetails.xls") as excel:
    df = pd.read_excel(excel, header=None)
    name = df.iloc[0, 6]  # G1单元格
    id_num = df.iloc[0, 9]  # J1单元格
    grzhye = df[df[5].notna()].iloc[-1, 5]  # F列最后一个非空值

    # 处理缴存记录数据
    jc_records = []
    valid_types = ["汇缴", "补缴", "补缴往月", "少缴补缴", "差额补缴"]
    data_rows = df.iloc[5:]  # 从第6行开始

    # 筛选有效记录并按单位分组
    records = []
    for _, row in data_rows.iterrows():
        if row[1] in valid_types:  # B列为业务类型
            records.append({
                'company': row[7],  # H列为单位名称
                'date': str(row[9])[:4] + "." + str(row[9])[4:6]  # J列为时间，格式化为yyyy.mm
            })

    # 按单位分组并合并日期
    company_groups = {}
    for record in records:
        if record['company'] not in company_groups:
            company_groups[record['company']] = []
        company_groups[record['company']].append(record['date'])

    # 合并每个单位的连续日期
    for company, dates in company_groups.items():
        merged_dates = merge_consecutive_dates(list(set(dates)))  # 去重并合并
        for date_range in merged_dates:
            jc_records.append({
                'company': company,
                'date_range': date_range,
                'sort_key': float(date_range.split('-')[0].replace('.', ''))  # 提取最早时间作为排序依据
            })

    # 按时间排序记录 - 确保按缴存时间段的起始日期正确排序
    jc_records.sort(key=lambda x: float(x['date_range'].split('-')[0].replace('.', '')))

# 读取提取.xls文件
extract_records = []
extract_file_path = "./DATA/ExtractDetails.xls"
logger.info(f"Checking extract file at: {extract_file_path}")

if not os.path.exists(extract_file_path):
    logger.error(f"Extract file not found at {extract_file_path}")
else:
    logger.info(f"Found extract file, size: {os.path.getsize(extract_file_path)} bytes")
    with pd.ExcelFile(extract_file_path) as excel:
        extract_df = pd.read_excel(excel, header=None)
        logger.info(f"Extract file loaded with {len(extract_df)} rows")

        # 获取所有缴存时间段用于筛选提取记录
        deposit_ranges = [r['date_range'] for r in jc_records]
        logger.info(f"Processing extract records against {len(deposit_ranges)} deposit ranges")

        extract_records = process_extract_records(extract_df, deposit_ranges)
        logger.info(f"Found {len(extract_records)} extract records to process")

# 创建新Excel文件
new_wb = Workbook()
new_ws = new_wb.active

# 写入基础数据
new_ws['B5'] = name  # 写入姓名到B5
new_ws['C5'] = id_num  # 写入证件号到C5
new_ws['D5'] = grzhye  # 写入账户余额到D5

# 写入缴存记录和对应的提取记录
current_row = 5
for record in jc_records:
    # 写入缴存记录(只写一次)
    new_ws[f'E{current_row}'] = record['date_range']  # 写入时间段到E列
    new_ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    new_ws[f'F{current_row}'] = record['company']  # 写入单位到F列
    new_ws[f'F{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # 写入对应的提取记录
    matching_extracts = [e for e in extract_records if e['deposit_range'] == record['date_range']]
    if matching_extracts:
        # 合并缴存记录单元格
        end_row = current_row + len(matching_extracts) - 1
        if end_row > current_row:
            new_ws.merge_cells(f'E{current_row}:E{end_row}')
            new_ws.merge_cells(f'F{current_row}:F{end_row}')

        for ext in matching_extracts:
            new_ws[f'I{current_row}'] = ext['date_range']  # 提取时间段
            new_ws[f'I{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            new_ws[f'J{current_row}'] = ext['amount']  # 提取金额
            new_ws[f'J{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            new_ws[f'K{current_row}'] = ext['type']  # 提取类型
            new_ws[f'K{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
    else:
        # 即使没有提取记录也要写入缴存记录
        new_ws[f'I{current_row}'] = ""
        new_ws[f'J{current_row}'] = ""
        new_ws[f'K{current_row}'] = ""
        current_row += 1

# 合并单元格并设置居中
last_row = current_row - 1
if last_row > 5:
    # 合并姓名列(B列)
    new_ws.merge_cells(f'B5:B{last_row}')
    new_ws['B5'].alignment = Alignment(horizontal='center', vertical='center')

    # 合并序号列(A列)
    new_ws.merge_cells(f'A5:A{last_row}')
    new_ws['A5'].alignment = Alignment(horizontal='center', vertical='center')

    # 合并证件号码列(C列)
    new_ws.merge_cells(f'C5:C{last_row}')
    new_ws['C5'].alignment = Alignment(horizontal='center', vertical='center')

    # 合并账户余额列(D列)
    new_ws.merge_cells(f'D5:D{last_row}')
    new_ws['D5'].alignment = Alignment(horizontal='center', vertical='center')

    # 合并当前缴存基数列(G列)
    new_ws.merge_cells(f'G5:G{last_row}')
    new_ws['G5'].alignment = Alignment(horizontal='center', vertical='center')

    # 合并当前月缴存额列(H列)
    new_ws.merge_cells(f'H5:H{last_row}')
    new_ws['H5'].alignment = Alignment(horizontal='center', vertical='center')

# 生成固定样式 start------------------------------------------------------------------------

new_ws.merge_cells('A1:K1')
from openpyxl.styles import Alignment, Font

new_ws['A1'].value = "附件1"
new_ws['A1'].font = Font(name='宋体', size=20, bold=True)
new_ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

new_ws.merge_cells('A2:K2')
from openpyxl.styles import Alignment, Font

new_ws['A2'].value = "住房公积金基本信息"
new_ws['A2'].font = Font(name='宋体', size=20, bold=True)
new_ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells('A3:A4')
from openpyxl.styles import Alignment, Font

new_ws['A3'].value = "序号"
new_ws['A3'].font = Font(name='宋体', size=11, bold=True)
new_ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells('B3:B4')
from openpyxl.styles import Alignment, Font

new_ws['B3'].value = "姓名"
new_ws['B3'].font = Font(name='宋体', size=11, bold=True)
new_ws['B3'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells('C3:C4')
from openpyxl.styles import Alignment, Font

new_ws['C3'].value = "证件号码"
new_ws['C3'].font = Font(name='宋体', size=11, bold=True)
new_ws['C3'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells('D3:D4')
from openpyxl.styles import Alignment, Font

new_ws['D3'].value = "账户余额"
new_ws['D3'].font = Font(name='宋体', size=11, bold=True)
new_ws['D3'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells('E3:H3')
from openpyxl.styles import Alignment, Font

new_ws['E3'].value = "缴存记录"
new_ws['E3'].font = Font(name='宋体', size=11, bold=True)
new_ws['E3'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['E4'].value = "缴存时间段"
new_ws['E4'].font = Font(name='宋体', size=11, bold=True)
new_ws['E4'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['F4'].value = "公积金缴存单位"
new_ws['F4'].font = Font(name='宋体', size=11, bold=True)
new_ws['F4'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['G4'].value = "当前缴存基数"
new_ws['G4'].font = Font(name='宋体', size=11, bold=True)
new_ws['G4'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['H4'].value = "当前月缴存额"
new_ws['H4'].font = Font(name='宋体', size=11, bold=True)
new_ws['H4'].alignment = Alignment(horizontal='center', vertical='center')

new_ws.merge_cells('I3:K3')
from openpyxl.styles import Alignment, Font

new_ws['I3'].value = "提取记录"
new_ws['I3'].font = Font(name='宋体', size=11, bold=True)
new_ws['I3'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['I4'].value = "提取时间"
new_ws['I4'].font = Font(name='宋体', size=11, bold=True)
new_ws['I4'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['J4'].value = "提取金额"
new_ws['J4'].font = Font(name='宋体', size=11, bold=True)
new_ws['J4'].alignment = Alignment(horizontal='center', vertical='center')

new_ws['K4'].value = "提取类型"
new_ws['K4'].font = Font(name='宋体', size=11, bold=True)
new_ws['K4'].alignment = Alignment(horizontal='center', vertical='center')

# new_ws.merge_cells('A6:K6')
# from openpyxl.styles import Alignment, Font
# new_ws['A6'].value = "查询截止时间为YYYY年MM月DD日HH时MI分"
# new_ws['A6'].font = Font(name='宋体', size=12)
# new_ws['A6'].alignment = Alignment(horizontal='center', vertical='center')


# 设置列宽
new_ws.column_dimensions['A'].width = 13
new_ws.column_dimensions['B'].width = 13
new_ws.column_dimensions['C'].width = 20
new_ws.column_dimensions['D'].width = 13
new_ws.column_dimensions['E'].width = 20
new_ws.column_dimensions['F'].width = 35
new_ws.column_dimensions['G'].width = 13
new_ws.column_dimensions['H'].width = 13
new_ws.column_dimensions['I'].width = 20
new_ws.column_dimensions['J'].width = 13
new_ws.column_dimensions['K'].width = 13

# 生成固定样式 end------------------------------------------------------------------------

# 添加查询截止时间行
last_row = new_ws.max_row + 1
new_ws.merge_cells(f'A{last_row}:K{last_row}')
new_ws[f'A{last_row}'].value = "查询截止时间为YYYY年MM月DD日HH时MI分"
new_ws[f'A{last_row}'].font = Font(name='宋体', size=12)
new_ws[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')

# 添加边框
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# 为所有单元格添加边框
for row in new_ws.iter_rows(min_row=1, max_row=new_ws.max_row):
    for cell in row:
        cell.border = thin_border

# 确保DATA目录存在
data_dir = "./DATA"
if not os.path.exists(data_dir):
    os.makedirs(data_dir)
    logger.info(f"Created directory: {data_dir}")

# 保存文件 - 处理文件被锁定的情况
output_path = os.path.join(data_dir, "output.xlsx")
logger.info(f"Attempting to save to: {output_path}")

try:
    # 尝试删除旧文件（如果存在且未被锁定）
    try:
        if os.path.exists(output_path):
            logger.info("Found existing output file, attempting to remove...")
            os.remove(output_path)
            logger.info("Successfully removed old file")
    except PermissionError as pe:
        logger.error(f"Error: {output_path} is locked. Please close the file and try again")
        logger.error(f"Details: {str(pe)}")
        exit(1)
    except Exception as e:
        logger.error(f"Error removing old file: {str(e)}")
        exit(1)

    # 保存新文件
    logger.info("Saving new file...")
    new_wb.save(output_path)
    logger.info(f"Success: File saved to {output_path}")

    # 验证文件是否创建成功
    if os.path.exists(output_path):
        logger.info(f"Verification: {output_path} exists, size: {os.path.getsize(output_path)} bytes")
    else:
        logger.error("Error: File save verification failed")
        input("按任意键退出...")
        exit(1)

    # 显示日志文件位置
    print("\n" + "=" * 50)
    print(f"日志文件位置: {log_file}")
    print("=" * 50 + "\n")
    if os.path.exists(log_file):
        print(f"日志文件已成功创建，大小: {os.path.getsize(log_file)} 字节")
    else:
        print("警告: 日志文件未生成")

    # 在EXE模式下保持窗口不关闭
    if getattr(sys, 'frozen', False):
        print("\n程序执行完成，按任意键退出...")
        input()
except Exception as e:
    logger.error(f"Error saving file: {str(e)}")
    logger.error(f"Full error details: {repr(e)}")
    input("按任意键退出...")
    exit(1)
